"""
SAFE Risk Controls Impact Script
--------------------------------
Reads Risk_Assessment and Settings from the .xlsm, then:
1) Calculates per-control decrease points: sum(grade * coverage)
   where grade = Severity * Probability.
2) Marks Applied = YES if control appears in current controls.
3) Computes ImpactPctOfGap = percent of the gap-to-appetite.
4) Prints results and writes to a NEW COPY of the workbook.
"""

import os
import re
import shutil
import warnings
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Silence openpyxl noise (NOT errors)
warnings.filterwarnings("ignore", message=".*extension is not supported.*", category=UserWarning)

# Config
FILE_PATH = r"G:\Shared drives\CTO\AI\AI Risk Assessment\Risk Assesment AI Readiness\Risk Assesment AI Readiness (version 1).xlsm"
RISK_SHEET = "Risk_Assessment"
SETTINGS_SHEET = "Settings"

OUTPUT_SHEET = "Controls impact"
OUTPUT_SUFFIX = "__output"  # creates a COPY, never touches original

REQUIRED_RISK_COLUMNS = [
    "severity",
    "probability",
    "security controls",
    "future planned controls",
]


def normalize_text(value) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def split_controls(cell_value) -> List[str]:
    if pd.isna(cell_value):
        return []
    return [part.strip() for part in re.split(r",|\n", str(cell_value)) if part.strip()]


def parse_pct_to_fraction(value) -> float:
    """Convert coverage to a fraction in [0, 1]. Accepts '10%', 10, 0.1, '0.1'."""
    if pd.isna(value):
        return np.nan

    if isinstance(value, str):
        s = value.strip()
        if s.endswith("%"):
            num = pd.to_numeric(s[:-1].strip(), errors="coerce")
            return np.nan if pd.isna(num) else float(num) / 100.0
        num = pd.to_numeric(s, errors="coerce")
        value = num

    if pd.isna(value):
        return np.nan

    v = float(value)
    return v if 0 <= v <= 1 else v / 100.0


def extract_controls_coverage(settings_path: str) -> Dict[str, float]:
    """Find the 'Controls'/'Coverage' table in Settings."""
    grid = pd.read_excel(settings_path, sheet_name=SETTINGS_SHEET, header=None, engine="openpyxl")
    lower = grid.astype(str).apply(lambda c: c.str.strip().str.lower())

    # Find adjacent headers: "controls" and "coverage"
    positions = zip(*lower.eq("controls").to_numpy().nonzero())
    for r, c in positions:
        if c + 1 < grid.shape[1] and lower.iat[r, c + 1] == "coverage":
            table = grid.iloc[r + 1 :, [c, c + 1]].copy()
            table.columns = ["Control", "Coverage_raw"]

            # Stop when Control is blank
            stop = table["Control"].isna() | (table["Control"].astype(str).str.strip() == "")
            if stop.any():
                table = table.loc[: stop.idxmax() - 1]

            table["Control"] = table["Control"].astype(str).str.strip()
            table["Coverage"] = table["Coverage_raw"].apply(parse_pct_to_fraction)
            table = table.dropna(subset=["Control", "Coverage"])

            return dict(zip(table["Control"].map(normalize_text), table["Coverage"]))

    raise RuntimeError("Couldn't locate the Controls/Coverage table in Settings sheet.")


def extract_risk_appetite_threshold(settings_path: str) -> float:
    """Find 'Risk Appetite Threshold' and return the numeric value next to it."""
    grid = pd.read_excel(settings_path, sheet_name=SETTINGS_SHEET, header=None, engine="openpyxl")
    lower = grid.astype(str).apply(lambda c: c.str.strip().str.lower())

    mask = lower.apply(lambda c: c.str.contains("risk appetite threshold", na=False))
    if not mask.any().any():
        raise RuntimeError("Couldn't find 'Risk Appetite Threshold' in Settings sheet.")

    r, c = next(zip(*mask.to_numpy().nonzero()))
    raw = grid.iat[r, c + 1] if c + 1 < grid.shape[1] else None
    val = pd.to_numeric(raw, errors="coerce")
    if pd.isna(val):
        raise RuntimeError(f"Risk Appetite Threshold value is not numeric: {raw}")

    return float(val)


def write_to_new_xlsm_copy(src_path: str, df: pd.DataFrame, sheet_name: str) -> str:
    """Write a sheet into a new .xlsm copy, keeping VBA intact."""
    base, ext = os.path.splitext(src_path)
    out_path = f"{base}{OUTPUT_SUFFIX}{ext}"

    shutil.copy2(src_path, out_path)

    wb = load_workbook(out_path, keep_vba=True)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(out_path)

    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        mode="a",
        engine_kwargs={"keep_vba": True},
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    return out_path


def validate_risk_columns(df: pd.DataFrame) -> Dict[str, str]:
    """Return a case-insensitive column map and validate required columns."""
    cols = {c.lower(): c for c in df.columns}
    missing = [c for c in REQUIRED_RISK_COLUMNS if c not in cols]
    if missing:
        raise RuntimeError(
            f"Missing required columns in Risk_Assessment: {missing}. Found: {list(df.columns)}"
        )
    return cols


def compute_control_impact(
    df: pd.DataFrame, coverage_map: Dict[str, float], risk_appetite: float
) -> pd.DataFrame:
    cols = validate_risk_columns(df)
    severity_col = cols["severity"]
    prob_col = cols["probability"]
    current_col = cols["security controls"]
    future_col = cols["future planned controls"]

    df = df.copy()
    df["grade"] = (
        pd.to_numeric(df[severity_col], errors="coerce").fillna(0)
        * pd.to_numeric(df[prob_col], errors="coerce").fillna(0)
    )

    total_risk = float(df["grade"].sum())
    gap_to_appetite = max(total_risk - risk_appetite, 0.0)

    df["controls_current"] = df[current_col].apply(split_controls)
    df["controls_future"] = df[future_col].apply(split_controls)

    curr = df.explode("controls_current").dropna(subset=["controls_current"]).copy()
    curr["control"] = curr["controls_current"]
    curr["control_norm"] = curr["control"].map(normalize_text)
    curr["applied"] = True

    fut = df.explode("controls_future").dropna(subset=["controls_future"]).copy()
    fut["control"] = fut["controls_future"]
    fut["control_norm"] = fut["control"].map(normalize_text)
    fut["applied"] = False

    exploded = pd.concat([curr, fut], ignore_index=True)
    exploded = exploded[exploded["control_norm"].isin(coverage_map)].copy()

    exploded["coverage"] = exploded["control_norm"].map(coverage_map)
    exploded["decrease_points"] = exploded["grade"] * exploded["coverage"]

    result = (
        exploded.groupby("control_norm", as_index=False)
        .agg(
            Control=("control", "first"),
            Applied=("applied", "max"),
            RisksCount=("grade", "size"),
            TotalDecreasePoints=("decrease_points", "sum"),
        )
        .sort_values("TotalDecreasePoints", ascending=False)
        .reset_index(drop=True)
    )

    result["Applied"] = result["Applied"].map({True: "YES", False: "NO"})
    result["ImpactPctOfGap"] = result["TotalDecreasePoints"].apply(
        lambda x: 0.0 if gap_to_appetite == 0 else (float(x) / gap_to_appetite) * 100.0
    )
    result.attrs["total_risk"] = total_risk
    result.attrs["gap_to_appetite"] = gap_to_appetite
    return result


def print_summary(
    out_df: pd.DataFrame, risk_appetite: float, total_risk: float, gap_to_appetite: float
) -> None:
    print_df = out_df.copy()
    print_df["TotalDecreasePoints"] = print_df["TotalDecreasePoints"].round(2)
    print_df["ImpactPctOfGap"] = print_df["ImpactPctOfGap"].map(lambda v: f"{v:.2f}%")

    print("\n=== Controls Impact ===")
    print(f"Risk Appetite Threshold: {risk_appetite}")
    print(f"Total Risk (sum grade): {total_risk}")
    print(f"Gap to appetite: {gap_to_appetite}")
    print()
    print(print_df.to_string(index=False))


def main() -> None:
    coverage_map = extract_controls_coverage(FILE_PATH)
    risk_appetite = extract_risk_appetite_threshold(FILE_PATH)
    df = pd.read_excel(FILE_PATH, sheet_name=RISK_SHEET, engine="openpyxl")

    result = compute_control_impact(df, coverage_map, risk_appetite)
    out_df = result[
        ["Control", "Applied", "RisksCount", "TotalDecreasePoints", "ImpactPctOfGap"]
    ].copy()

    print_summary(
        out_df,
        risk_appetite=risk_appetite,
        total_risk=result.attrs["total_risk"],
        gap_to_appetite=result.attrs["gap_to_appetite"],
    )

    out_file = write_to_new_xlsm_copy(FILE_PATH, out_df, OUTPUT_SHEET)
    print("\nDONE. Output written to:", out_file)


if __name__ == "__main__":
    main()
