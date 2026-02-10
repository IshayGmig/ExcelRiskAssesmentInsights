# ===============================
# SAFE Risk Controls Impact Script
# ===============================
# What it does:
# 1) Reads Risk_Assessment + Settings from your .xlsm
# 2) Calculates per-control risk decrease points: SUM(grade * coverage)
#    where grade = Severity * Probability
#    and the control is counted if it appears in:
#       - "Security Controls" (current)
#       - "Future Planned Controls" (planned)
# 3) Adds:
#       - Applied (YES if appears anywhere in "Security Controls", else NO)
#       - ImpactPctOfGap (percent of the gap-to-appetite covered by that control)
# 4) Prints to terminal (with points + %)
# 5) Writes to a NEW COPY of the workbook (never overwrites original)
#    in a sheet named "Controls impact"

import os
import re
import shutil
import warnings
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# -------------------------------
# Silence openpyxl noise (NOT errors)
# -------------------------------
warnings.filterwarnings("ignore", message=".*extension is not supported.*", category=UserWarning)

# -------------------------------
# CONFIG
# -------------------------------
FILE_PATH = r"G:\Shared drives\CTO\AI\AI Risk Assessment\Risk Assesment AI Readiness\Risk Assesment AI Readiness (version 1).xlsm"
RISK_SHEET = "Risk_Assessment"
SETTINGS_SHEET = "Settings"

OUTPUT_SHEET = "Controls impact"
OUTPUT_SUFFIX = "__output"  # creates a COPY, never touches original

# -------------------------------
# Helpers
# -------------------------------
def norm(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    return re.sub(r"\s+", " ", str(x).strip().lower())


def split_controls(cell) -> List[str]:
    if pd.isna(cell):
        return []
    return [p.strip() for p in re.split(r",|\n", str(cell)) if p.strip()]


def parse_pct_to_fraction(x) -> float:
    """Convert coverage to a fraction in [0,1]. Accepts '10%', 10, 0.1, '0.1'."""
    if pd.isna(x):
        return np.nan

    if isinstance(x, str):
        s = x.strip()
        if s.endswith("%"):
            num = pd.to_numeric(s[:-1].strip(), errors="coerce")
            return np.nan if pd.isna(num) else float(num) / 100.0
        num = pd.to_numeric(s, errors="coerce")
        x = num

    if pd.isna(x):
        return np.nan

    v = float(x)
    return v if 0 <= v <= 1 else v / 100.0


# -------------------------------
# Settings extraction
# -------------------------------
def extract_controls_coverage(settings_path: str) -> Dict[str, float]:
    """
    Finds the 'Controls'/'Coverage' table in Settings and returns:
      { normalized_control_name: coverage_fraction }
    """
    grid = pd.read_excel(settings_path, sheet_name=SETTINGS_SHEET, header=None, engine="openpyxl")
    lower = grid.astype(str).apply(lambda c: c.str.strip().str.lower())

    # Find adjacent headers: "controls" and "coverage"
    positions = zip(*lower.eq("controls").to_numpy().nonzero())
    for r, c in positions:
        if c + 1 < grid.shape[1] and lower.iat[r, c + 1] == "coverage":
            table = grid.iloc[r + 1 :, [c, c + 1]].copy()
            table.columns = ["Control", "Coverage_raw"]

            # Stop when Control is blank
            stop = table["Control"].isna() | (table["Control"].astype(str).str.strip() == "")
            if stop.any():
                table = table.loc[: stop.idxmax() - 1]

            table["Control"] = table["Control"].astype(str).str.strip()
            table["Coverage"] = table["Coverage_raw"].apply(parse_pct_to_fraction)
            table = table.dropna(subset=["Control", "Coverage"])

            return dict(zip(table["Control"].map(norm), table["Coverage"]))

    raise RuntimeError("Couldn't locate the Controls/Coverage table in Settings sheet.")


def extract_risk_appetite_threshold(settings_path: str) -> float:
    """
    Finds 'Risk Appetite Threshold' in Settings and returns the numeric value from the next column.
    """
    grid = pd.read_excel(settings_path, sheet_name=SETTINGS_SHEET, header=None, engine="openpyxl")
    lower = grid.astype(str).apply(lambda c: c.str.strip().str.lower())

    mask = lower.apply(lambda c: c.str.contains("risk appetite threshold", na=False))
    if not mask.any().any():
        raise RuntimeError("Couldn't find 'Risk Appetite Threshold' in Settings sheet.")

    r, c = next(zip(*mask.to_numpy().nonzero()))
    raw = grid.iat[r, c + 1] if c + 1 < grid.shape[1] else None
    val = pd.to_numeric(raw, errors="coerce")
    if pd.isna(val):
        raise RuntimeError(f"Risk Appetite Threshold value is not numeric: {raw}")

    return float(val)


# -------------------------------
# SAFE writer: writes to a COPY (.xlsm preserved)
# -------------------------------
def write_to_new_xlsm_copy(src_path: str, df: pd.DataFrame, sheet_name: str) -> str:
    base, ext = os.path.splitext(src_path)
    out_path = f"{base}{OUTPUT_SUFFIX}{ext}"

    shutil.copy2(src_path, out_path)

    wb = load_workbook(out_path, keep_vba=True)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(out_path)

    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        mode="a",
        engine_kwargs={"keep_vba": True},
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    return out_path


# -------------------------------
# MAIN
# -------------------------------
def main() -> None:
    coverage_map = extract_controls_coverage(FILE_PATH)
    risk_appetite = extract_risk_appetite_threshold(FILE_PATH)

    df = pd.read_excel(FILE_PATH, sheet_name=RISK_SHEET, engine="openpyxl")

    # Required columns (case-insensitive)
    cols = {c.lower(): c for c in df.columns}
    required = ["severity", "probability", "security controls", "future planned controls"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise RuntimeError(f"Missing required columns in Risk_Assessment: {missing}. Found: {list(df.columns)}")

    severity_col = cols["severity"]
    prob_col = cols["probability"]
    current_col = cols["security controls"]
    future_col = cols["future planned controls"]

    # Grade per risk
    df["grade"] = (
        pd.to_numeric(df[severity_col], errors="coerce").fillna(0)
        * pd.to_numeric(df[prob_col], errors="coerce").fillna(0)
    )

    total_risk = float(df["grade"].sum())
    gap_to_appetite = max(total_risk - risk_appetite, 0.0)

    # Split controls per row
    df["controls_current"] = df[current_col].apply(split_controls)
    df["controls_future"] = df[future_col].apply(split_controls)

    # Explode current + future separately so we can compute Applied flag properly
    curr = df.explode("controls_current").dropna(subset=["controls_current"]).copy()
    curr["control"] = curr["controls_current"]
    curr["control_norm"] = curr["control"].map(norm)
    curr["applied"] = True

    fut = df.explode("controls_future").dropna(subset=["controls_future"]).copy()
    fut["control"] = fut["controls_future"]
    fut["control_norm"] = fut["control"].map(norm)
    fut["applied"] = False

    exploded = pd.concat([curr, fut], ignore_index=True)

    # Keep only controls that exist in Settings table
    exploded = exploded[exploded["control_norm"].isin(coverage_map)].copy()

    exploded["coverage"] = exploded["control_norm"].map(coverage_map)
    exploded["decrease_points"] = exploded["grade"] * exploded["coverage"]

    # Aggregate per control
    result = (
        exploded.groupby("control_norm", as_index=False)
        .agg(
            Control=("control", "first"),
            Applied=("applied", "max"),  # True if it appears at least once in current controls
            RisksCount=("grade", "size"),
            TotalDecreasePoints=("decrease_points", "sum"),
        )
        .sort_values("TotalDecreasePoints", ascending=False)
        .reset_index(drop=True)
    )

    result["Applied"] = result["Applied"].map({True: "YES", False: "NO"})

    # Percent column: % of (total_risk - risk_appetite)
    result["ImpactPctOfGap"] = result["TotalDecreasePoints"].apply(
        lambda x: 0.0 if gap_to_appetite == 0 else (float(x) / gap_to_appetite) * 100.0
    )

    # Output columns you want
    out_df = result[["Control", "Applied", "RisksCount", "TotalDecreasePoints", "ImpactPctOfGap"]].copy()

    # Print to terminal nicely: points + percent with %
    print_df = out_df.copy()
    print_df["TotalDecreasePoints"] = print_df["TotalDecreasePoints"].round(2)
    print_df["ImpactPctOfGap"] = print_df["ImpactPctOfGap"].map(lambda v: f"{v:.2f}%")

    print("\n=== Controls Impact ===")
    print(f"Risk Appetite Threshold: {risk_appetite}")
    print(f"Total Risk (sum grade): {total_risk}")
    print(f"Gap to appetite: {gap_to_appetite}")
    print()
    print(print_df.to_string(index=False))

    # Write to NEW COPY of the workbook (safe)
    out_file = write_to_new_xlsm_copy(FILE_PATH, out_df, OUTPUT_SHEET)
    print("\nDONE. Output written to:", out_file)


if __name__ == "__main__":
    main()
